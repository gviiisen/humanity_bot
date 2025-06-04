import traceback
import threading
import os
import time
import random
from openpyxl import load_workbook, Workbook
from datetime import datetime
from collections import deque
from collections import defaultdict
import logging

# 存储不同文件的锁
file_locks = defaultdict(threading.Lock)

wallet_lock = threading.Lock()  # 创建线程锁

def ensure_directory_exists(path):
    directory = os.path.dirname(path)
    if not os.path.exists(directory):
        os.makedirs(directory)


def write_csv(path, data):
    """
    将数据写入CSV文件
    
    Args:
        path: 文件路径
        data: 要写入的数据（列表或元组）
    """
    try:
        # 检查数据类型
        if not isinstance(data, (list, tuple)):
            raise ValueError(f"数据必须是列表或元组格式，当前格式为: {type(data)}")

        normalized_path = os.path.normpath(path)
        ensure_directory_exists(normalized_path)

        # 获取文件锁
        if normalized_path not in file_locks:
            file_locks[normalized_path] = threading.Lock()
        lock = file_locks[normalized_path]

        with lock:
            existing_lines = set()
            # 读取现有内容
            if os.path.exists(normalized_path):
                try:
                    with open(normalized_path, mode='r', encoding='utf-8') as file:
                        existing_lines = set(line.strip() for line in file)
                except UnicodeDecodeError:
                    logging.error(f"读取文件 {normalized_path} 时出现编码错误")
                    existing_lines = set()

            # 把新数据转换成字符串
            try:
                new_line = '----'.join(str(item) for item in data)
            except Exception as e:
                raise ValueError(f"数据转换失败: {str(e)}")

            if new_line in existing_lines:
                logging.info(f"数据已存在，不重复添加: {data}")
                return

            # 追加新数据
            try:
                with open(normalized_path, mode='a', encoding='utf-8', newline='') as file:
                    file.write(new_line + '\n')
                logging.info(f"成功写入数据到文件 {normalized_path}")
            except Exception as e:
                raise IOError(f"写入文件失败: {str(e)}")

    except Exception as e:
        logging.error(f"写入CSV文件时出错: {str(e)}")
        traceback.print_exc()
        raise


def get_today_log_filename(data_dir):
    """根据当前日期生成日志文件名，并检查文件是否存在，如果不存在则创建一个空文件"""
    today = datetime.now().strftime("%Y%m%d")  # 获取当前日期，格式为YYYYMMDD
    log_filename = f"exe_{today}.log"
    # 如果 data 文件夹不存在，就创建它
    os.makedirs(data_dir, exist_ok=True)
    log_path = os.path.join(data_dir, log_filename)

    # 检查文件是否存在，如果不存在则创建一个空的日志文件
    if not os.path.exists(log_path):
        with open(log_path, 'w') as file:
            pass  # 创建一个空文件

    return log_path


def get_today_fail_log_filename(data_dir):
    """根据当前日期生成Excel文件名，并检查文件是否存在，如果不存在则创建一个空Excel文件"""
    today = datetime.now().strftime("%Y%m%d")  # 获取当前日期，格式为YYYYMMDD
    excel_filename = f"check_{today}.xlsx"  # 使用 .xlsx 扩展名
    # 如果 data 文件夹不存在，就创建它
    os.makedirs(data_dir, exist_ok=True)
    excel_path = os.path.join(data_dir, excel_filename)

    # 检查文件是否存在，如果不存在则创建一个空的Excel文件
    if not os.path.exists(excel_path):
        # 创建一个新的Workbook对象
        wb = Workbook()
        # 保存到指定路径
        wb.save(excel_path)

    return excel_path


def append_to_excel(excel_path, contain_id, num):
    """
    向指定Excel文件追加数据（线程安全）：
    - A列写入contain_id
    - B列写入num
    - 从第3行开始追加写入
    """
    with excel_lock:  # 使用锁确保线程安全
        # 如果文件不存在，创建一个新文件并初始化前两行
        if not os.path.exists(excel_path):
            wb = Workbook()
            ws = wb.active
            # 设置表头（第1行）
            ws['A1'] = 'Contain ID'
            ws['B1'] = 'Number'
            # 保存初始文件
            wb.save(excel_path)
        else:
            # 加载现有Excel文件
            wb = load_workbook(excel_path)
            ws = wb.active

        # 找到当前最后一行（从第3行开始追加）
        max_row = max(2, ws.max_row)  # 确保从第3行开始，即使已有数据
        if max_row < 2:
            max_row = 2  # 如果文件是空的，至少从第3行开始

        # 写入新数据到下一行
        next_row = max_row + 1
        ws[f'A{next_row}'] = contain_id  # A列写入contain_id
        ws[f'B{next_row}'] = num  # B列写入num

        # 保存文件
        wb.save(excel_path)
        wb.close()


def load_data_from_txt(file_path):
    """读取txt文件，将每行作为一个元素装入集合以去重"""
    data_set = set()
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            for line in file:
                data_set.add(line.strip())  # 去掉换行符并添加到集合中
    except Exception as e:
        traceback.print_exc()
        raise Exception(f"加载文件 {file_path} 失败")
    return data_set


def load_data_from_txt(file_path):
    """读取txt文件，将每行作为一个元素装入集合以去重"""
    data_set = set()
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            for line in file:
                data_set.add(line.strip())  # 去掉换行符并添加到集合中
    except Exception as e:
        traceback.print_exc()
        raise Exception(f"加载文件 {file_path} 失败")
    return data_set

def load_data_from_txt(file_path):
    """读取txt文件，将去重后的每行数据装入 deque"""
    wallet_deque = deque()
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            unique_lines = set(line.strip() for line in file if line.strip())  # 去重 & 去空行
        wallet_deque.extend(unique_lines)  # 一次性装入 deque
    except Exception as e:
        raise Exception(f"加载文件 {file_path} 失败: {e}")
    return wallet_deque



def get_random_element(data_set):
    """从集合中随机获取一个元素"""
    if data_set:
        return random.choice(list(data_set))
    else:
        raise ValueError("集合为空，无法获取随机元素")

def update_wallet_txt(file_path, old_mnemonic, new_mnemonic, wallet_deque):
    """
    线程安全地更新 wallet.txt：
    1. 删除旧助记词
    2. 追加新助记词到文件
    3. 将新助记词放入 deque 左侧
    """
    with wallet_lock:  # 加锁，防止多个线程同时写入
        try:
            # 读取现有助记词
            with open(file_path, 'r', encoding='utf-8') as file:
                mnemonics = [line.strip() for line in file.readlines()]

            # 删除旧助记词
            mnemonics = [mn for mn in mnemonics if mn != old_mnemonic]

            # 追加新助记词
            mnemonics.append(new_mnemonic)

            # 覆盖写入文件
            with open(file_path, 'w', encoding='utf-8') as file:
                file.write('\n'.join(mnemonics) + '\n')

            # 更新 deque，把新助记词放到左边
            wallet_deque.appendleft(new_mnemonic)

        except Exception as e:
            print(f"更新 wallet.txt 失败: {e}")
